out_dir = '/fs/clip-quiz/nbalepur/QG_vs_QA/data/'

from openpyxl import Workbook
from openpyxl.styles import Alignment
import datasets
ds = datasets.load_dataset('nbalepur/QG_vs_QA')['subset']
cats = set(ds['category'])

wb = Workbook()
for c in cats:
    wb.create_sheet(c)
    wb.active = wb[c]
    ws = wb.active
    ds_ = ds.filter(lambda ex: ex['category'] == c)
    e, t = ds_['entity'], ds_['context']

    for idx in range(len(e)):
        ws[f'A{idx*2+2}'] = e[idx]

        if c == 'num_text':
            ws[f'C{idx*2+2}'] = t[idx]

wb.save(f"{out_dir}/quality_check.xlsx")